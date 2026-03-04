import argparse
import re
import unicodedata
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


BOOL_MAP = {
    "yes": 1,
    "y": 1,
    "true": 1,
    "t": 1,
    "1": 1,
    "no": 0,
    "n": 0,
    "false": 0,
    "f": 0,
    "0": 0,
}

PLACEHOLDER_PATTERNS = (
    r"^\(enter.*\)$",
    r"^\(verify.*\)$",
)

SECTION_RE = re.compile(r"^(\d+)\)\s+(.+)$")
KEY_VALUE_RE = re.compile(r"^([^:]+):\s*(.+)$")
SEPARATOR_RE = re.compile(r"^[-=]+$")


def slugify(text: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", ascii_text.strip().lower())
    return slug.strip("_") or "col"


def normalize_text(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = re.sub(r"\s+", " ", value).strip()
        return cleaned or None
    return value


def normalize_placeholder(value: object) -> object:
    cleaned = normalize_text(value)
    if not isinstance(cleaned, str):
        return cleaned

    for pattern in PLACEHOLDER_PATTERNS:
        if re.match(pattern, cleaned, flags=re.IGNORECASE):
            return None
    return cleaned


def coerce_boolean_object_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if out[col].dtype != "object":
            continue
        lowered = out[col].dropna().astype(str).str.strip().str.lower()
        if lowered.empty:
            continue
        unique_vals = set(lowered.unique().tolist())
        if unique_vals.issubset(set(BOOL_MAP.keys())):
            out[col] = (
                out[col]
                .astype("string")
                .str.strip()
                .str.lower()
                .map(BOOL_MAP)
                .astype("Int64")
            )
    return out


def normalize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    mapping = {}
    used = set()
    for col in df.columns:
        new_col = slugify(str(col))
        base = new_col
        i = 2
        while new_col in used:
            new_col = f"{base}_{i}"
            i += 1
        used.add(new_col)
        mapping[col] = new_col
    return df.rename(columns=mapping)


def clean_string_cells(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if out[col].dtype == "object" or str(out[col].dtype).startswith("string"):
            out[col] = out[col].map(normalize_text)
    return out


def write_csv(df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(output_path, index=False)


def process_my_bom_csv(input_path: Path, output_dir: Path) -> tuple[Path, int, int]:
    df = pd.read_csv(input_path)
    df = normalize_column_names(df)
    df = clean_string_cells(df)
    df = coerce_boolean_object_columns(df)

    numeric_candidates = ["mass_g", "recycled_content_pct", "contains_hazardous_substance"]
    for col in numeric_candidates:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    output_path = output_dir / f"{input_path.stem}_synthcity_ready.csv"
    write_csv(df, output_path)
    return output_path, len(df), len(df.columns)


def find_header_row(ws) -> int:
    for row_idx in range(1, min(30, ws.max_row) + 1):
        values = [normalize_text(ws.cell(row_idx, col_idx).value) for col_idx in range(1, min(ws.max_column, 25) + 1)]
        lowered = [str(v).lower() for v in values if v is not None]
        joined = " | ".join(lowered)
        if "position number" in joined and "component number" in joined:
            return row_idx
    raise ValueError(f"Could not find BOM header row in sheet '{ws.title}'.")


def normalize_excel_scalar(value: object, force_string: bool) -> object:
    cleaned = normalize_text(value)
    if cleaned is None:
        return None
    if force_string:
        return str(cleaned)
    return cleaned


def process_bom_xlsx(input_path: Path, output_dir: Path) -> tuple[Path, int, int, pd.DataFrame]:
    wb = load_workbook(input_path, data_only=True)
    if "BOM_ENG" not in wb.sheetnames:
        raise ValueError(f"{input_path.name} missing 'BOM_ENG' sheet.")

    ws = wb["BOM_ENG"]
    header_row = find_header_row(ws)

    raw_headers = [normalize_text(ws.cell(header_row, col_idx).value) for col_idx in range(1, ws.max_column + 1)]
    headers_with_pos = [(idx + 1, h) for idx, h in enumerate(raw_headers) if h]
    if not headers_with_pos:
        raise ValueError(f"{input_path.name}: no usable headers found in BOM_ENG.")

    records = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        record = {}
        row_has_value = False
        for col_idx, header in headers_with_pos:
            value = ws.cell(row_idx, col_idx).value
            if value is not None and str(value).strip() != "":
                row_has_value = True
            record[header] = value
        if row_has_value:
            records.append(record)

    if not records:
        raise ValueError(f"{input_path.name}: BOM_ENG has no data rows after header.")

    df = pd.DataFrame(records)
    df = normalize_column_names(df)

    # Keep IDs/codes as strings to preserve formatting (e.g., leading zeros).
    force_string_cols = {
        "resolution",
        "position_number",
        "assembly_indicator",
        "component_number",
        "base_amount",
        "compensation_unit",
        "vi_material",
    }
    for col in df.columns:
        force_string = col in force_string_cols or col.endswith("_number")
        df[col] = df[col].map(lambda v: normalize_excel_scalar(v, force_string=force_string))

    if "componentmng_kme" in df.columns:
        df["componentmng_kme"] = pd.to_numeric(df["componentmng_kme"], errors="coerce")
    if "vi_recommendation" in df.columns:
        df["vi_recommendation"] = pd.to_numeric(df["vi_recommendation"], errors="coerce")

    df = df.dropna(axis=1, how="all")
    df = df.dropna(axis=0, how="all")
    df = clean_string_cells(df)

    product_scope = slugify(input_path.stem.replace("_BOM", ""))
    df["source_product_scope"] = product_scope
    df["source_sheet"] = "BOM_ENG"
    df["source_file"] = input_path.name

    output_path = output_dir / f"{slugify(input_path.stem)}_synthcity_ready.csv"
    write_csv(df, output_path)
    return output_path, len(df), len(df.columns), df.copy()


def append_with_delimiter(record: dict[str, object], key: str, value: str, delim: str = " || ") -> None:
    if key not in record or record[key] in (None, ""):
        record[key] = value
    else:
        record[key] = f"{record[key]}{delim}{value}"


def parse_dpp_text_file(input_path: Path) -> dict[str, object]:
    record: dict[str, object] = {
        "source_file": input_path.name,
        "source_type": "dpp_txt",
    }

    current_section = "document_header"
    with input_path.open("r", encoding="utf-8") as f:
        for raw_line in f:
            line = normalize_text(raw_line)
            if not line:
                continue
            if line == "END OF DOCUMENT" or "END OF DOCUMENT" in line:
                continue
            if SEPARATOR_RE.match(line):
                continue

            section_match = SECTION_RE.match(line)
            if section_match:
                current_section = slugify(section_match.group(2))
                continue

            if line.startswith("IMPORTANT NOTE"):
                current_section = "important_note"
                continue

            if line.startswith("DIGITAL PRODUCT PASSPORT"):
                record["document_title"] = line
                continue

            if line.startswith("•"):
                bullet_key = f"{current_section}__bullet_items"
                append_with_delimiter(record, bullet_key, line.lstrip("•").strip())
                continue

            kv = KEY_VALUE_RE.match(line)
            if kv:
                raw_key = slugify(kv.group(1))
                value = normalize_placeholder(kv.group(2))
                full_key = f"{current_section}__{raw_key}"
                record[full_key] = value
                continue

            notes_key = f"{current_section}__notes"
            append_with_delimiter(record, notes_key, line)

    return record


def process_lexmark_txt(input_path: Path, output_dir: Path) -> tuple[Path, int, int, pd.DataFrame]:
    record = parse_dpp_text_file(input_path)
    df = pd.DataFrame([record])
    df = normalize_column_names(df)
    df = clean_string_cells(df)
    df = coerce_boolean_object_columns(df)
    df = df.dropna(axis=1, how="all")

    output_path = output_dir / f"{input_path.stem}_synthcity_ready.csv"
    write_csv(df, output_path)
    return output_path, len(df), len(df.columns), df.copy()


def build_dataframe_union(records: Iterable[dict[str, object]]) -> pd.DataFrame:
    df = pd.DataFrame(list(records))
    if df.empty:
        return df
    df = normalize_column_names(df)
    df = clean_string_cells(df)
    df = df.dropna(axis=1, how="all")
    return df


def main() -> None:
    ap = argparse.ArgumentParser(description="Create SynthCity-ready CSVs from mixed DPP/BOM files.")
    ap.add_argument("--input-dir", default="data", help="Directory containing source files.")
    ap.add_argument("--output-dir", default="data/synthcity_ready", help="Directory to write normalized CSV files.")
    args = ap.parse_args()

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    summary_rows = []
    lexmark_frames = []
    vitocal_frames = []

    csv_path = input_dir / "my_bom_dpp.csv"
    if csv_path.exists():
        out, rows, cols = process_my_bom_csv(csv_path, output_dir)
        summary_rows.append(
            {
                "source_file": csv_path.name,
                "source_type": "csv",
                "status": "converted",
                "rows": rows,
                "columns": cols,
                "output_file": str(out),
            }
        )

    for xlsx_path in sorted(input_dir.glob("*.xlsx")):
        out, rows, cols, bom_df = process_bom_xlsx(xlsx_path, output_dir)
        summary_rows.append(
            {
                "source_file": xlsx_path.name,
                "source_type": "xlsx_bom",
                "status": "converted",
                "rows": rows,
                "columns": cols,
                "output_file": str(out),
            }
        )
        vitocal_frames.append(bom_df)

    for txt_path in sorted(input_dir.glob("lexmark_*_dpp_full_*.txt")):
        out, rows, cols, lexmark_df = process_lexmark_txt(txt_path, output_dir)
        summary_rows.append(
            {
                "source_file": txt_path.name,
                "source_type": "txt_dpp",
                "status": "converted",
                "rows": rows,
                "columns": cols,
                "output_file": str(out),
            }
        )
        lexmark_frames.append(lexmark_df)

    if lexmark_frames:
        lexmark_df = pd.concat(lexmark_frames, ignore_index=True, sort=False)
        lexmark_df = normalize_column_names(lexmark_df)
        lexmark_df = clean_string_cells(lexmark_df)
        lexmark_df = lexmark_df.dropna(axis=1, how="all")
        lexmark_output = output_dir / "lexmark_dpp_combined_synthcity_ready.csv"
        write_csv(lexmark_df, lexmark_output)
        summary_rows.append(
            {
                "source_file": "lexmark_*_dpp_full_*.txt",
                "source_type": "txt_dpp_combined",
                "status": "combined",
                "rows": len(lexmark_df),
                "columns": len(lexmark_df.columns),
                "output_file": str(lexmark_output),
            }
        )

    if vitocal_frames:
        vitocal_df = pd.concat(vitocal_frames, ignore_index=True, sort=False)
        vitocal_output = output_dir / "vitocal_bom_combined_synthcity_ready.csv"
        write_csv(vitocal_df, vitocal_output)
        summary_rows.append(
            {
                "source_file": "*.xlsx::BOM_ENG",
                "source_type": "xlsx_bom_combined",
                "status": "combined",
                "rows": len(vitocal_df),
                "columns": len(vitocal_df.columns),
                "output_file": str(vitocal_output),
            }
        )

    summary_df = pd.DataFrame(summary_rows)
    summary_output = output_dir / "conversion_summary.csv"
    write_csv(summary_df, summary_output)
    print(summary_df.to_string(index=False))
    print(f"\nSummary written to: {summary_output}")


if __name__ == "__main__":
    main()
